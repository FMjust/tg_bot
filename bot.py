import matplotlib
matplotlib.use('Agg')
import os
import telebot
import threading
import time
import traceback
from datetime import datetime
from telebot import types
from smc import detect_liquidity_sweep_and_bos, detect_po3_amd_model
from binance_api import get_candles
from fvg import find_fvg_zones as find_fvg_zones_fvg  # если понадобится отдельно, иначе можно убрать
from liquidity import find_swing_highs_lows, is_near_liquidity
from momentum import calculate_rsi
from chart import plot_signal_chart
from logger import log_signal_to_csv
from po3 import analyze_po3_structure
from orderflow import OrderFlowAnalyzer
import concurrent.futures
from volume import average_volume, is_volume_spike, find_volume_clusters, filter_zones_by_volume, zone_has_cluster
import numpy as np
import openpyxl
from openpyxl import load_workbook
from topdown import (
    get_trend_htf, get_pd_zones, filter_signals_by_htf,
    find_poi_zones, find_fta_zones
)
ORDERFLOW_SYMBOLS = ["BTCUSDT", "ETHUSDT", "SOLUSDT", "OPUSDT", "SNXUSDT", "BNBUSDT", "XRPUSDT", "DOGEUSDT", "ADAUSDT", "TRXUSDT",
    "SUIUSDT", "LINKUSDT", "AVAXUSDT", "XLMUSDT", "SHIBUSDT", "BCHUSDT", "HBARUSDT", "XMRUSDT", "TONUSDT", "LTCUSDT",
    "PEPEUSDT", "UNIUSDT", "AAVEUSDT", "TAOUSDT", "NEARUSDT", "APTUSDT", "ONDOUSDT", "ICPUSDT", "ETCUSDT",
    "TRUMPUSDT", "RENDERUSDT", "POLUSDT", "VETUSDT", "ENAUSDT", "FETUSDT", "WLDUSDT", "ARBUSDT", "ALGOUSDT",
    "ATOMUSDT", "FILUSDT", "JUPUSDT", "TIAUSDT", "INJUSDT", "STXUSDT", "SUSDT", "SEIUSDT", "IMXUSDT", "QNTUSDT",
    "WIFUSDT", "FORMUSDT", "GRTUSDT", "DEXEUSDT", "CRVUSDT", "MKRUSDT", "JASMYUSDT", "ZECUSDT", "GALAUSDT",
    "THETAUSDT", "CAKEUSDT", "PENGUUSDT", "ENSUSDT", "PAXGUSDT", "IOTAUSDT", "SANDUSDT", "LDOUSDT", "PYTHUSDT",
    "PENDLEUSDT"]
orderflow_analyzers = {}
for symbol in ORDERFLOW_SYMBOLS:
    analyzer = OrderFlowAnalyzer(symbol)
    analyzer.start()
    orderflow_analyzers[symbol] = analyzer

def log_transaction_to_excel(symbol, direction, entry, sl, tp, result, strategy, filename="transactions.xlsx"):
    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Time", "Symbol", "Direction", "Entry", "SL", "TP", "Result", "Strategy"])
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([now, symbol, direction, entry, sl, tp, result, strategy])
        wb.save(filename)
    except Exception as e:
        print(f"[ExcelLog] Error logging transaction: {e}")

BOT_TOKEN = "8133935884:AAGlKSU1qZAk4mCqMsi7UbyUAWRp5h_Dvj1"
bot = telebot.TeleBot(BOT_TOKEN)

symbols = [
    "BTCUSDT", "ETHUSDT", "SOLUSDT", "OPUSDT", "SNXUSDT", "BNBUSDT", "XRPUSDT", "DOGEUSDT", "ADAUSDT", "TRXUSDT",
    "SUIUSDT", "LINKUSDT", "AVAXUSDT", "XLMUSDT", "SHIBUSDT", "BCHUSDT", "HBARUSDT", "XMRUSDT", "TONUSDT", "LTCUSDT",
    "PEPEUSDT", "UNIUSDT", "AAVEUSDT", "TAOUSDT", "NEARUSDT", "APTUSDT", "ONDOUSDT", "ICPUSDT", "ETCUSDT",
    "TRUMPUSDT", "RENDERUSDT", "POLUSDT", "VETUSDT", "ENAUSDT", "FETUSDT", "WLDUSDT", "ARBUSDT", "ALGOUSDT",
    "ATOMUSDT", "FILUSDT", "JUPUSDT", "TIAUSDT", "INJUSDT", "STXUSDT", "SUSDT", "SEIUSDT", "IMXUSDT", "QNTUSDT",
    "WIFUSDT", "FORMUSDT", "GRTUSDT", "DEXEUSDT", "CRVUSDT", "MKRUSDT", "JASMYUSDT", "ZECUSDT", "GALAUSDT",
    "THETAUSDT", "CAKEUSDT", "PENGUUSDT", "ENSUSDT", "PAXGUSDT", "IOTAUSDT", "SANDUSDT", "LDOUSDT", "PYTHUSDT",
    "PENDLEUSDT"
]

monitoring_enabled = True
chat_id = None
active_signals = {}

WEIGHTS = {
    "ob": 2,
    "liquidity": 2.5,
    "multi_bos": 1,
    "fvg": 1,
    "rsi": 1
}

MIN_SIGNAL_STRENGTH = 0

def main_menu():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row("🔍 Скан")
    markup.row("📊 Статус", "🧠 Стратегии")
    markup.row("▶️ Старт", "⏸️ Пауза")
    markup.row("🧹 Очистить")
    return markup

def fetch_all_candles(symbol):
    return {
        "h1": get_candles(symbol, interval="1h", limit=50),
        "m15": get_candles(symbol, interval="15m", limit=50),
        "d1": get_candles(symbol, interval="1d", limit=50),
    }

def get_trend(candles):
    if not candles or len(candles) < 2:
        return "neutral"
    return "up" if candles[-1]["close"] > candles[0]["close"] else "down"

def is_flat_market(candles, flat_threshold=0.005):
    if not candles or len(candles) < 2:
        return False
    closes = [c['close'] for c in candles]
    high = max(closes)
    low = min(closes)
    avg = sum(closes) / len(closes)
    return (high - low) / avg < flat_threshold

def near_key_levels(entry, swing_levels, key_threshold=0.002):
    if not swing_levels:
        return False
    for level in swing_levels:
        if abs(level - entry) / entry < key_threshold:
            return True
    return False

def calculate_ema(prices, period=50):
    prices = np.array(prices)
    if len(prices) < period:
        return None
    weights = np.exp(np.linspace(-1., 0., period))
    weights /= weights.sum()
    a = np.convolve(prices, weights, mode='valid')
    return float(a[-1]) if len(a) > 0 else None

def calculate_sma(prices, period=50):
    prices = np.array(prices)
    if len(prices) < period:
        return None
    return float(np.mean(prices[-period:]))

def fvg_retest_strategy(candles):
    if not candles or len(candles) < 50:
        return {}

    zones = find_fvg_zones_fvg(candles)
    last_close = candles[-1]['close']
    last_open = candles[-1]['open']
    closes = [c['close'] for c in candles]

    ema50 = calculate_ema(closes, period=50)

    signal = None

    for zone in zones:
        zone_low, zone_high = zone['from'], zone['to']
        if zone_low <= last_close <= zone_high:
            direction = 'long' if last_close > last_open else 'short'
            rsi = calculate_rsi(candles)
            trend = get_trend(candles)
            trend_match = ((trend == "up" and direction == "long") or (trend == "down" and direction == "short"))
            rsi_match = (rsi > 50 and direction == "long") or (rsi < 50 and direction == "short")

            if ema50 is not None:
                if direction == "long" and last_close < ema50:
                    break
                if direction == "short" and last_close > ema50:
                    break

            if trend_match and rsi_match:
                signal = {
                    "signal": True,
                    "direction": direction,
                    "entry": last_close,
                    "sl": last_close * (0.99 if direction == "long" else 1.01),
                    "tp": last_close * (1.02 if direction == "long" else 0.98),
                    "reason": f"Цена вернулась в FVG + EMA50 фильтр + подтверждение RSI и тренд",
                    "strategy": "FVG Re-Test",
                    "ob": False
                }
            break
    return signal or {}

def calculate_risk_reward(entry, sl, tp, direction):
    if direction == "long":
        risk = entry - sl
        reward = tp - entry
    else:
        risk = sl - entry
        reward = entry - tp
    if risk <= 0:
        return None
    return round(reward / risk, 2)

def evaluate_signal_strength(signal, candles_h1, symbol, candles_m15=None, candles_d1=None):
    required_keys = ("entry", "sl", "tp", "direction")
    if not signal or not signal.get("signal") or not all(k in signal for k in required_keys):
        return 0, ["Недостаточно данных для оценки сигнала"], False, False, False, False, 0
    score = 0
    reasons = []

    rsi = calculate_rsi(candles_h1)
    if (signal['direction'] == 'long' and rsi > 70) or (signal['direction'] == 'short' and rsi < 30):
        reasons.append(f"RSI {'> 70' if signal['direction']=='long' else '< 30'} — no trade")
        return 0, reasons, False, False, False, False, rsi

    if is_flat_market(candles_h1):
        reasons.append("Flat market — no trade")
        return 0, reasons, False, False, False, False, rsi

    swing_levels = sum(find_swing_highs_lows(candles_d1 or get_candles(symbol, "1d", 50)), [])
    if near_key_levels(signal['entry'], swing_levels):
        reasons.append("Entry near key level — avoid trade")
        return 0, reasons, False, False, False, False, rsi

    if signal.get("ob"):
        score += WEIGHTS["ob"]
    else:
        reasons.append("OB не подтверждён")

    fvg_ok = any(
        zone['from'] <= signal['entry'] <= zone['to']
        for zone in find_fvg_zones_fvg(candles_h1)
    )
    if fvg_ok:
        score += WEIGHTS["fvg"]
    else:
        reasons.append("Нет зоны FVG")

    multi_bos = detect_liquidity_sweep_and_bos(candles_m15 or get_candles(symbol, "15m", 50)).get("signal")
    if multi_bos:
        score += WEIGHTS["multi_bos"]
    else:
        reasons.append("Нет подтверждения BOS на M15")

    near_liq = is_near_liquidity(
        signal['entry'],
        swing_levels
    )
    if near_liq:
        score += WEIGHTS["liquidity"]
    else:
        reasons.append("Вдали от зоны ликвидности")

    rsi_ok = (
        (signal['direction'] == 'long' and rsi > 50)
        or (signal['direction'] == 'short' and rsi < 50)
    )
    if rsi_ok:
        score += WEIGHTS["rsi"]
    else:
        reasons.append("RSI против сигнала")

    m15_trend = get_trend(candles_m15) if candles_m15 else get_trend(get_candles(symbol, "15m", 50))
    main_trend = get_trend(candles_h1)
    m15_confirm = ((m15_trend == main_trend) and
                   ((main_trend == "up" and signal['direction'] == 'long') or
                    (main_trend == "down" and signal['direction'] == 'short')))
    if not m15_confirm:
        reasons.append("Нет подтверждения M15 тренда/структуры")
        return 0, reasons, fvg_ok, multi_bos, near_liq, rsi_ok, rsi

    return score, reasons, fvg_ok, multi_bos, near_liq, rsi_ok, rsi
    
def scan_symbol(symbol, candles_dict):
    candles_h1 = candles_dict["h1"]
    candles_m15 = candles_dict["m15"]
    candles_d1 = candles_dict["d1"]

    results = []

    # --- Top-Down анализ и Premium/Discount ---
    candles_htf = candles_d1  # D1 - старший ТФ
    htf_trend = get_trend_htf(candles_htf)
    premium_zone, discount_zone = get_pd_zones(candles_htf)

    # Определяем зоны для POI (entry) и FTA (take)
    if htf_trend == 'long':
        poi_zone = discount_zone
        fta_zone = premium_zone
    else:
        poi_zone = premium_zone
        fta_zone = discount_zone

    # Поиск POI и FTA
    pois = find_poi_zones(candles_h1, poi_zone, htf_trend)
    ftas = find_fta_zones(candles_h1, htf_trend, fta_zone)

    signals = [
        detect_liquidity_sweep_and_bos(candles_h1),
        detect_po3_amd_model(candles_h1),
        analyze_po3_structure(candles_h1),
        fvg_retest_strategy(candles_h1)
    ]
    # Фильтрация сигналов по тренду HTF
    filtered_signals = filter_signals_by_htf(signals, htf_trend)

    last_candle = candles_h1[-1]
    avg_vol = average_volume(candles_h1)
    clusters = find_volume_clusters(candles_h1, coef=2.0, n=40)

    # --- Order Flow анализатор (получаем заранее, используем внутри цикла) ---
    analyzer = orderflow_analyzers.get(symbol.upper())
    orderflow_info = analyzer.get_info() if analyzer else None

    for signal in filtered_signals:
        required_keys = ("entry", "sl", "tp", "direction")
        if not signal or not signal.get("signal") or not all(k in signal for k in required_keys):
            continue

        in_poi = any(poi['from'] <= signal['entry'] <= poi['to'] for poi in pois)
        if not in_poi:
            continue

        # ОБЪЁМНЫЙ ФИЛЬТР
        if not is_volume_spike(last_candle, candles_h1, coef=1.5):
            reasons = ["Нет всплеска объёма на последней свече"]
            results.append({
                "signal": signal, "score": 0, "reasons": reasons,
                "fvg_ok": False, "multi_bos": False,
                "near_liq": False, "rsi_ok": False, "rsi": None
            })
            continue

        # ФИЛЬТР ПО КЛАСТЕРАМ В POI
        pois_with_cluster = [z for z in pois if zone_has_cluster(z, clusters)]
        if not pois_with_cluster:
            reasons = ["Нет объёмного кластера в зонах POI"]
            results.append({
                "signal": signal, "score": 0, "reasons": reasons,
                "fvg_ok": False, "multi_bos": False,
                "near_liq": False, "rsi_ok": False, "rsi": None
            })
            continue
        pois = pois_with_cluster

        # --- Order Flow фильтр ---
        if orderflow_info:
            if orderflow_info['cvd'] < 0:
                reasons = ["Order Flow: доминируют продажи, сигнал фильтруется"]
                results.append({
                    "signal": signal, "score": 0, "reasons": reasons,
                    "fvg_ok": False, "multi_bos": False,
                    "near_liq": False, "rsi_ok": False, "rsi": None
                })
                continue

        # --- Оценка силы сигнала ---
        score, reasons, fvg_ok, multi_bos, near_liq, rsi_ok, rsi = evaluate_signal_strength(
            signal, candles_h1, symbol,
            candles_m15=candles_m15, candles_d1=candles_d1
        )

        entry = signal.get("entry", 0)
        if isinstance(entry, (float, int)):
            if premium_zone[0] <= entry <= premium_zone[1]:
                reasons.append("Entry в зоне Premium (верхняя половина диапазона)")
            elif discount_zone[0] <= entry <= discount_zone[1]:
                reasons.append("Entry в зоне Discount (нижняя половина диапазона)")
            else:
                reasons.append("Entry вне зон Premium/Discount")

        if ftas:
            signal['tp'] = ftas[0]['from']

        if pois:
            poi_strings = [f"{p['type']}({p['from']:.2f}-{p['to']:.2f})" for p in pois]
            reasons.append(f"POI зоны: {poi_strings}")
        if ftas:
            reasons.append(f"FTA зона для TP: {ftas[0]['from']:.2f}-{ftas[0]['to']:.2f}")

        # ДОБАВЛЯЕМ ОБЪЁМНУЮ ИНФОРМАЦИЮ
        reasons.append(f"Объём: {last_candle['volume']} (средний: {int(avg_vol)})")
        reasons.append(f"Всплеск объёма: {'Да' if is_volume_spike(last_candle, candles_h1) else 'Нет'}")
        reasons.append(f"Кластеров за 40 свечей: {len(clusters)}")

        # ДОБАВЛЯЕМ Order Flow инфу (если есть)
        if orderflow_info:
            reasons.append(
                f"Order Flow: CVD={orderflow_info['cvd']:.2f}, "
                f"Buy={orderflow_info['buy_vol']:.2f}, Sell={orderflow_info['sell_vol']:.2f}"
            )

        results.append({
            "signal": signal, "score": score, "reasons": reasons,
            "fvg_ok": fvg_ok, "multi_bos": multi_bos,
            "near_liq": near_liq, "rsi_ok": rsi_ok, "rsi": rsi
        })

    return symbol, results
def detailed_manual_analysis(symbol, chat_id):
    candles_dict = fetch_all_candles(symbol)
    candles_h1 = candles_dict["h1"]
    candles_m15 = candles_dict["m15"]
    candles_d1 = candles_dict["d1"]

    strategies = [
        ("Liquidity Sweep + BOS", detect_liquidity_sweep_and_bos),
        ("PO3 + AMD Model", detect_po3_amd_model),
        ("FVG Re-Test", fvg_retest_strategy)
    ]

    text = f"📊 Детальный анализ для {symbol} (H1):\n"
    found_any = False
    required_keys = ("entry", "sl", "tp", "direction")

    for strat_name, strat_func in strategies:
        signal = strat_func(candles_h1)
        text += f"\n<b>{strat_name}</b>\n"
        if not (signal and signal.get("signal")):
            text += "⛔ Нет сигнала. "
            if signal and signal.get("reason"):
                text += signal.get("reason") + "\n"
            else:
                text += "Условия стратегии не выполнены или недостаточно данных.\n"
            continue
        if not all(k in signal for k in required_keys):
            text += "⛔ Нет сигнала. Причина: стратегия не вернула все параметры (entry, sl, tp, direction).\n"
            continue

        found_any = True
        score, reasons, fvg_ok, multi_bos, near_liq, rsi_ok, rsi = evaluate_signal_strength(
            signal, candles_h1, symbol, candles_m15, candles_d1
        )
        rr = calculate_risk_reward(
            signal['entry'],
            signal['sl'],
            signal['tp'],
            signal['direction']
        )
        direction_emoji = "🟢 LONG" if signal['direction'] == 'long' else "🔴 SHORT"
        text += (
            f"{direction_emoji}\n"
            f"🎯 Entry: {signal['entry']:.2f}\n"
            f"⛔ SL: {signal['sl']:.2f}\n"
            f"✅ TP: {signal['tp']:.2f}\n"
            f"📈 R|R: {rr if rr is not None else 'N/A'}\n"
            f"💡 Причина: {signal.get('reason')}\n"
            f"RSI: {rsi:.2f}\n"
            f"Сила сигнала: {score} из 7.5\n"
        )
        # --- NEW: Log signal transaction when generated ---
        log_transaction_to_excel(
            symbol=symbol,
            direction=signal['direction'],
            entry=signal['entry'],
            sl=signal['sl'],
            tp=signal['tp'],
            result="signal",
            strategy=strat_name
        )
        if reasons:
            text += "\n".join([f"❕ {r}" for r in reasons if r]) + "\n"
    if not found_any:
        text += "\n⚠️ Нет активных сигналов по стратегиям."
    os.makedirs("graphic", exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
    chart_filename = f"graphic/{timestamp}_{symbol}.png"
    plot_signal_chart(candles_h1, {}, symbol, chart_filename)
    with open(chart_filename, "rb") as img:
        bot.send_photo(chat_id, img, caption=text, parse_mode="HTML")

def handle_scan_all(message):
    found_signals = 0
    results_text = ""
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        candles_map = dict(executor.map(
            lambda symbol: (symbol, fetch_all_candles(symbol)),
            symbols
        ))

    for symbol in symbols:
        _, results = scan_symbol(symbol, candles_map[symbol])
        if results:
            for res in results:
                signal = res["signal"]
                score = res["score"]
                required_keys = ("entry", "sl", "tp", "direction")
                if score < MIN_SIGNAL_STRENGTH or not all(k in signal for k in required_keys):
                    continue
                rr = calculate_risk_reward(
                    signal['entry'],
                    signal['sl'],
                    signal['tp'],
                    signal['direction']
                )
                direction_emoji = "🟢 LONG" if signal['direction'] == 'long' else "🔴 SHORT"
                results_text += (
                    f"\n<b>{symbol}</b> | <b>{signal.get('strategy', 'Strategy')}</b>\n"
                    f"{direction_emoji}\n"
                    f"🎯 Entry: {signal['entry']:.2f} | SL: {signal['sl']:.2f} | TP: {signal['tp']:.2f} | R|R: {rr if rr is not None else 'N/A'}\n"
                    f"💡 Причина: {signal.get('reason')}\n"
                    f"Сила сигнала: {score} из 7.5\n"
                )
                found_signals += 1
                # --- NEW: Log signal transaction when found in scan ---
                log_transaction_to_excel(
                    symbol=symbol,
                    direction=signal['direction'],
                    entry=signal['entry'],
                    sl=signal['sl'],
                    tp=signal['tp'],
                    result="signal",
                    strategy=signal.get('strategy', 'Strategy')
                )
                os.makedirs("graphic", exist_ok=True)
                timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
                chart_filename = f"graphic/{timestamp}_{symbol}_{signal.get('strategy', 'strat')}.png"
                plot_signal_chart(candles_map[symbol]["h1"], signal, symbol, chart_filename)
                with open(chart_filename, "rb") as img:
                    bot.send_photo(message.chat.id, img)

    if found_signals == 0:
        bot.send_message(message.chat.id, "ℹ️ Сигналы не найдены.", parse_mode="HTML")
    else:
        bot.send_message(message.chat.id, f"✅ Найдено сигналов: {found_signals}\n{results_text}", parse_mode="HTML")

@bot.message_handler(commands=['strategies'])
def handle_strategies(message):
    text = (
    "📘 Активные стратегии бота:\n"
    "1️⃣ Liquidity Sweep + BOS\n"
    "2️⃣ PO3 + AMD Model\n"
    "3️⃣ PO3 (Sessions)\n"               
    "4️⃣ FVG Re-Test"
)
    bot.send_message(message.chat.id, text)

@bot.message_handler(commands=['start'])
def handle_start(message):
    global chat_id
    chat_id = message.chat.id
    bot.send_message(message.chat.id,
        "Привет! Я TWBINTRADEBOT — торговый помощник по Smart Money стратегиям.\n"
        "Я автоматически слежу за рынком и присылаю сигналы.\n"
        "Выберите действие через меню ниже.",
        reply_markup=main_menu()
    )

@bot.message_handler(func=lambda message: message.text == "🔍 Скан")
def handle_scan_btn(message):
    bot.send_message(message.chat.id, "📡 Поиск сигналов по всем инструментам...")
    handle_scan_all(message)

@bot.message_handler(func=lambda message: message.text == "📊 Статус")
def handle_status(message):
    status_text = "🟢 Автоанализ: ВКЛЮЧЕН" if monitoring_enabled else "🔴 Автоанализ: ВЫКЛЮЧЕН"
    if not active_signals:
        bot.send_message(message.chat.id, f"{status_text}\nℹ️ Нет активных сигналов.")
        return

    response = f"{status_text}\n📊 Активные сигналы:\n"
    for (symbol, strategy), sig in active_signals.items():
        response += (
            f"\n📍 {symbol} ({strategy})\n"
            f"🔹 Направление: {'LONG' if sig['direction'] == 'long' else 'SHORT'}\n"
            f"🎯 Entry: {sig['entry']:.2f}\n"
            f"⛔ SL: {sig['sl']:.2f}\n"
            f"✅ TP: {sig['tp']:.2f}\n"
        )
    bot.send_message(message.chat.id, response)

@bot.message_handler(func=lambda message: message.text == "🧠 Стратегии")
def handle_strategies_btn(message):
    handle_strategies(message)

@bot.message_handler(func=lambda message: message.text == "▶️ Старт")
def handle_start_btn(message):
    handle_start(message)

@bot.message_handler(func=lambda message: message.text == "⏸️ Пауза")
def handle_pause(message):
    global monitoring_enabled
    monitoring_enabled = False
    bot.send_message(message.chat.id, "⏸️ Автоматический анализ приостановлен.")

@bot.message_handler(func=lambda message: message.text == "🧹 Очистить")
def handle_clear(message):
    try:
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id - 1)
    except Exception as e:
        if "message to delete not found" not in str(e):
            bot.send_message(message.chat.id, f"Ошибка при удалении: {e}")


def calculate_smart_money_rr(entry, sl, tp, direction):
    try:
        risk = abs(entry - sl)
        reward = abs(tp - entry)
        rr_ratio = round(reward / risk, 2) if risk != 0 else 0
        return rr_ratio
    except Exception as e:
        print(f"[ERROR] Ошибка в расчете R|R: {e}")
        return 0

def send_startup_notification():
    global monitoring_enabled
    status_text = "🟢 Автоанализ включен" if monitoring_enabled else "🔴 Автоанализ выключен"
    if chat_id:
        bot.send_message(chat_id, f"🤖 Бот запущен.\n{status_text}\nДля помощи нажмите /start")

def monitor_signals():
    while True:
        if not monitoring_enabled or not chat_id:
            time.sleep(10)
            continue

        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            candles_map = dict(executor.map(
                lambda symbol: (symbol, fetch_all_candles(symbol)),
                symbols
            ))

        for symbol in symbols:
            _, results = scan_symbol(symbol, candles_map[symbol])
            for res in results:
                signal = res["signal"]
                score = res["score"]
                strategy = signal.get("strategy", "Strategy")
                key = (symbol, strategy)
                required_keys = ("entry", "sl", "tp", "direction")
                if score < MIN_SIGNAL_STRENGTH or not all(k in signal for k in required_keys):
                    continue
                if key in active_signals:
                    continue  
                active_signals[key] = signal
                rr = calculate_risk_reward(
                    signal['entry'],
                    signal['sl'],
                    signal['tp'],
                    signal['direction']
                )
                direction_emoji = "🟢 LONG" if signal['direction'] == 'long' else "🔴 SHORT"
                text = (
                    f"\n<b>{symbol}</b> | <b>{signal.get('strategy', 'Strategy')}</b>\n"
                    f"{direction_emoji}\n"
                    f"🎯 Entry: {signal['entry']:.2f} | SL: {signal['sl']:.2f} | TP: {signal['tp']:.2f} | R|R: {rr if rr is not None else 'N/A'}\n"
                    f"💡 Причина: {signal.get('reason')}\n"
                    f"Сила сигнала: {score} из 7.5\n"
                )
                log_transaction_to_excel(
                    symbol=symbol,
                    direction=signal['direction'],
                    entry=signal['entry'],
                    sl=signal['sl'],
                    tp=signal['tp'],
                    result="signal",
                    strategy=signal.get('strategy', 'Strategy')
                )
                os.makedirs("graphic", exist_ok=True)
                timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
                chart_filename = f"graphic/{timestamp}_{symbol}_{signal.get('strategy', 'strat')}.png"
                plot_signal_chart(candles_map[symbol]["h1"], signal, symbol, chart_filename)
                with open(chart_filename, "rb") as img:
                    bot.send_photo(chat_id, img, caption=text, parse_mode="HTML")
            

        for (symbol, strategy), signal in list(active_signals.items()):
            required_keys = ("entry", "sl", "tp", "direction")
            if not signal or not all(k in signal for k in required_keys):
                continue  

            candles = candles_map[symbol]["h1"]
            current_price = candles[-1]['close']
            rsi = calculate_rsi(candles)

            if signal['direction'] == 'long' and rsi < 50:
                bot.send_message(chat_id, f"⚠️ Сигнал по {symbol} ({strategy}) стал неактуальным из-за RSI")
                # --- NEW: Log outcome ---
                log_transaction_to_excel(
                    symbol=symbol,
                    direction=signal['direction'],
                    entry=signal['entry'],
                    sl=signal['sl'],
                    tp=signal['tp'],
                    result="inactive_RSI",
                    strategy=strategy
                )
                del active_signals[(symbol, strategy)]
                continue
            elif signal['direction'] == 'short' and rsi > 50:
                bot.send_message(chat_id, f"⚠️ Сигнал по {symbol} ({strategy}) стал неактуальным из-за RSI")
                log_transaction_to_excel(
                    symbol=symbol,
                    direction=signal['direction'],
                    entry=signal['entry'],
                    sl=signal['sl'],
                    tp=signal['tp'],
                    result="inactive_RSI",
                    strategy=strategy
                )
                del active_signals[(symbol, strategy)]
                continue

            if signal['direction'] == 'long':
                if current_price <= signal['sl']:
                    bot.send_message(chat_id, f"❌ Стоп-лосс достигнут по {symbol}, цена: {current_price:.2f}")
                    log_transaction_to_excel(
                        symbol=symbol,
                        direction=signal['direction'],
                        entry=signal['entry'],
                        sl=signal['sl'],
                        tp=signal['tp'],
                        result="SL",
                        strategy=strategy
                    )
                    del active_signals[(symbol, strategy)]
                elif current_price >= signal['tp']:
                    bot.send_message(chat_id, f"✅ Тейк-профит достигнут по {symbol}, цена: {current_price:.2f}")
                    log_transaction_to_excel(
                        symbol=symbol,
                        direction=signal['direction'],
                        entry=signal['entry'],
                        sl=signal['sl'],
                        tp=signal['tp'],
                        result="TP",
                        strategy=strategy
                    )
                    del active_signals[(symbol, strategy)]
            else:
                if current_price >= signal['sl']:
                    bot.send_message(chat_id, f"❌ Стоп-лосс достигнут по {symbol}, цена: {current_price:.2f}")
                    log_transaction_to_excel(
                        symbol=symbol,
                        direction=signal['direction'],
                        entry=signal['entry'],
                        sl=signal['sl'],
                        tp=signal['tp'],
                        result="SL",
                        strategy=strategy
                    )
                    del active_signals[(symbol, strategy)]
                elif current_price <= signal['tp']:
                    bot.send_message(chat_id, f"✅ Тейк-профит достигнут по {symbol}, цена: {current_price:.2f}")
                    log_transaction_to_excel(
                        symbol=symbol,
                        direction=signal['direction'],
                        entry=signal['entry'],
                        sl=signal['sl'],
                        tp=signal['tp'],
                        result="TP",
                        strategy=strategy
                    )
                    del active_signals[(symbol, strategy)]

        time.sleep(120)

threading.Thread(target=monitor_signals, daemon=True).start()
print("🚀 TWBINTRADEBOT is running.")
send_startup_notification()
while True:
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        print(f"Polling crashed with error: {e}. Restarting in 15 seconds...")
        traceback.print_exc()
        time.sleep(15)
